import vk_api
from vk_api.utils import get_random_id
from datetime import datetime, date
import time
import os
import sys
from openpyxl import Workbook, load_workbook

def write_message(sender, message):
	""" Отсылает личное сообщение """
	authorize.method('messages.send', {'user_id': sender, 'message': message, 'random_id': get_random_id()})
	

vk_token = ""

authorize = vk_api.VkApi(token = vk_token)

# открыли файл со списком подписчиков
wb = load_workbook('subscribers.xlsx')
ws = wb.active

sub_list = []

# читаем файл с постом
with open('newpost.txt') as np:
	newpost = np.read()
print(newpost)

q = input('\nRegions? ')

# если рассылка всем подписчикам
if q == 'all':
	for row in range(2,ws.max_row+1):
		#  выбираем всех подписчиков в список
		sub_list.append(str(ws.cell(row,2).value))
	print(sub_list)

# если отдельный регион
elif q != 'all' and len(q) == 2:
	for row in range(2,ws.max_row+1):
		if ws.cell(row,3).value == q:
			#  выбираем подписчиков по региону в список
			sub_list.append(str(ws.cell(row,2).value))
	print(sub_list)

# если группа регионов
elif q != 'all' and len(q) > 2:
	# собираем группу регионов в список
	q_list = q.split(' ')
	for sub in q_list:
		for row in range(2,ws.max_row+1):
			if ws.cell(row,3).value == sub:
				#  выбираем подписчиков по региону в список
				sub_list.append(str(ws.cell(row,2).value))
	print(sub_list)

# если подписчиков в этих регионах нет
if len(sub_list) == 0:
	print('\nNo subscribers.')
	sys.exit()


# *********** Рассылка ***********
print('Sending message...')
for subscriber in sub_list:
	time.sleep(0.35)
	write_message(subscriber, newpost)
print('End of sending.')
# *********** конец рассылки **********





