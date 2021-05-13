import vk_api
import requests
from datetime import datetime, date, time
import time
import os
import sys
from openpyxl import Workbook, load_workbook
from vk_api.longpoll import VkLongPoll, VkEventType
from vk_api.utils import get_random_id
from vk_api.keyboard import VkKeyboard, VkKeyboardColor


# Клавиатуры
keybo_da_net = VkKeyboard(one_time=True)
keybo_da_net.add_button('Да, всё верно', color=VkKeyboardColor.POSITIVE)
keybo_da_net.add_button('Нет, ошибка', color=VkKeyboardColor.NEGATIVE)

keybo_otpiska = VkKeyboard(one_time=True)
keybo_otpiska.add_button('Отписаться', color=VkKeyboardColor.NEGATIVE)

def write_message(sender, message, keyb):
	""" Отправка личного сообщения ВК с клавиатурой """
	if keyb == 2:
		authorize.method('messages.send', {'user_id': sender, 'message': message, 'random_id': get_random_id(), 'keyboard': keybo_otpiska.get_keyboard()})
	if keyb == 1:
		authorize.method('messages.send', {'user_id': sender, 'message': message, 'random_id': get_random_id(), 'keyboard': keybo_da_net.get_keyboard()})
	if keyb == 0:
		authorize.method('messages.send', {'user_id': sender, 'message': message, 'random_id': get_random_id()})

def get_name(sender):
	""" Возвращает имя отправителя """
	sender_info = getting_api.users.get(user_ids = sender)[0]
	name = sender_info.get('first_name')
	return name

def erlog(*args):
	""" Логирование ошибок в файл """
	l = open("vk_error.txt", 'a', encoding='windows-1251')
	print('Дата '+str(datetime.today())[:-10], file=l)
	print(*args, file=l)
	l.close()

def vklog(*args):
	""" Логирование работы бота в файл """
	l = open("vk_log.txt", 'a', encoding='utf-8')
	print('Дата '+str(datetime.today())[:-10], file=l)
	print(*args, file=l)
	l.close()


	# ****** TELEGRAM ******
t_vknmu = ""
t_chatid = ""

t_err = ""
t_err_chatid = ""
	# ******* конец TELEGRAM *******

	# ****** ВКОНТАКТЕ *******
token = ""
authorize = vk_api.VkApi(token = token)
longpoll = VkLongPoll(authorize)
getting_api = authorize.get_api()


regions = {'02': ' Республика Башкортостан',
		'12': ' Республика Марий Эл',
		'16': ' Республика Татарстан',
		'21': ' Республика Чувашия',
		'31': ' Белгородская область',
		'32': ' Брянская область',
		'44': ' Костромская область',
		'46': ' Курская область',
		'48': ' Липецкая область',
		'52': ' Нижегородская область',
		'54': ' Новосибирская область',
		'55': ' Омская область',
		'56': ' Оренбургская область',
		'57': ' Орловская область',
		'58': ' Пензенская область',
		'59': ' Пермский край',
		'63': ' Самарская область',
		'64': ' Саратовская область',
		'66': ' Свердловская область',
		'68': ' Тамбовская область',
		'72': ' Тюменская область (г. Тюмень)',
		'73': ' Ульяновская область',
		'74': ' Челябинская область',
		# '75': ' Забайкальский край',
		'82': ' Республика Крым',
		'97': ' Москва и Московская область',
		'98': ' Санкт-Петербург и Ленинградская область',
		}
		
senders = {}		# словарь писавших боту (в сообщество)
subscribers = {}	# словарь подписчиков рассылки


# база подписчиков
if not os.path.isfile('subscribers.xlsx'):
	wb = Workbook('subscribers.xlsx')
	ws = wb.create_sheet("Подписчики")
	ws.append(["date", "ID", "REG"])
	wb.save(os.path.dirname(os.path.abspath(__file__)) + "/subscribers.xlsx")
	wb = load_workbook('subscribers.xlsx')
else:
	wb = load_workbook('subscribers.xlsx')
ws = wb.active

# считываем эксель, создаем список подписчиков
for row in range(2,ws.max_row+1):
	if(ws.cell(row,2).value is not None):
		subscribers[ws.cell(row,2).value] = ws.cell(row,3).value

	# """ Цикл прослушивания """
while True:
	try:
		for event in longpoll.listen():
				time.sleep(0.35)
				
				# Боту прислали текст
				if event.type == VkEventType.MESSAGE_NEW and event.to_me and event.text:
					resieved_message = event.text
					# определяем id пользователя
					sender = event.user_id
					# Получаем имя пользователя
					name = get_name(sender)
					
					# Если пользователя нет в списке писавших - добавляем в список с шагом 0
					if sender not in senders:	
						senders[sender] = 0
						vklog('Пользователь '+str(sender)+' пишет: "'+resieved_message+'"')
											# """ ПРИВЕТСТВИЕ """
						write_message(sender, 'Привет, '+name+'! Я робот-помощник проекта "Прогнозы НМУ". &#129302;', 0)
						time.sleep(1)
		
							# """ АДМИН """
					if resieved_message == '/sub' and sender == 602267660:
						write_message(sender, 'Количество подписчиков: '+str(len(subscribers)), 0)
						write_message(sender, str(subscribers), 0)
						continue
			
					if resieved_message == '/stop' and sender == 602267660:
						write_message(sender, 'Бот остановлен '+str(datetime.today())[:-10], 0)
						sys.exit()
							# """  КОНЕЦ АДМИНКИ """
									
					if sender in subscribers and senders[sender] == 0:
						senders[sender] = 5
					
			
					# """ ОТПИСКА """
					if sender in senders and (senders[sender] == 6 or senders[sender] == 5):
						if resieved_message.lower() == 'отписаться':
							write_message(sender,'&#9940;Ваша подписка аннулирована, '+
								name+'.\nНо вы всегда можете её возобновить. Просто напишите мне, договоримся. &#128521;\nДо встречи! &#129302;', 0)
							
							stroka = 2
							while ws['B'+str(stroka)].value != sender:
								stroka +=1
							ws.delete_rows(stroka, 1)
							wb.save(os.path.dirname(os.path.abspath(__file__)) + "/subscribers.xlsx")
			
							vklog('Пользователь '+str(sender)+' отписался: регион '+regions[subscribers[sender]])
							# постим в ВК техканал
							r = requests.post('https://api.telegram.org/bot{0}/sendMessage?chat_id={1}&text={3}{2}{4}{2}{5}{6}'.format(
								t_vknmu, t_chatid, "\n",
								 '🔴Пользователь отписался!',
								'Регион '+subscribers[sender]+', '+regions[subscribers[sender]],
								'https://vk.com/id'+str(sender),
								'&parse_mode=Markdown&disable_web_page_preview=false'))
			
							del senders[sender]
							del subscribers[sender]
						
						else:
							vklog('Пользователь '+str(sender)+', ошибка отписки: "'+resieved_message+'"')
							write_message(sender, 'Я вас не понял.&#128530;', 0)
							time.sleep(2)
							senders[sender] = 5
			
			
					
					# """ РЕГИОН ПОДПИСКИ, запрос отписки """
					if sender in senders and senders[sender] == 5:
						vklog('Пользователь '+str(sender)+', подписан '+regions[subscribers[sender]]+' пишет: "'+resieved_message+'"')
						write_message(sender, name+', сейчас вы подписаны на уведомления о прогнозах НМУ по региону\n&#9925;&#9925;&#9925;\n'+
								regions[subscribers[sender]]+'\n&#9925;&#9925;&#9925;\nНадеюсь, вам нравится наш проект! Но если что - можно в любой момент отменить подписку, нажав красную кнопку. \nИли напишите "Отписаться"', 2)
						senders[sender] = 6
			
			
					# """ ПОДТВЕРЖДЕНИЕ ВЫБОРА """
					if sender in senders and senders[sender] == 2 and (resieved_message == 'Да, всё верно' or resieved_message.lower() == 'да'):
						# - добавляем в словарь подписчиков с номером региона
						subscribers[sender] = reg_sub
						ws.append([str(datetime.today())[:-10], sender, reg_sub])
						wb.save(os.path.dirname(os.path.abspath(__file__)) + "/subscribers.xlsx")
						write_message(sender, 'Отлично, '+name+'!&#128077;\n Ваша подписка будет активирована в течение часа.&#128337;\n'+
								'Уведомления об изменении прогноза НМУ на сайте Росгидромета будут приходить вам в мессенджер ВКонтакте не реже одного раза в сутки.\n'+
								'Если прогноз за сутки не изменится, придёт сообщение "Прогноз НМУ не изменился".\n'+
								'Удачной работы!&#128521;', 0)
						
						vklog('Пользователь '+str(sender)+' подписался, регион: '+regions[subscribers[sender]])
						# постим в ВК техканал
						r = requests.post('https://api.telegram.org/bot{0}/sendMessage?chat_id={1}&text={3}{2}{4}{2}{5}{6}'.format(
							t_vknmu, t_chatid, "\n",
							 '❇Пользователь подписался!',
							'Регион '+subscribers[sender]+', '+regions[subscribers[sender]],
							'https://vk.com/id'+str(sender),
							'&parse_mode=Markdown&disable_web_page_preview=false'))
			
						reg_sub = ''
						senders[sender] = 5
					
					elif sender in senders and senders[sender] == 2 and (resieved_message == 'Нет, ошибка' or resieved_message.lower() == 'нет'):
						vklog('Пользователь '+str(sender)+', ошибка выбора региона: "'+resieved_message+'"')
						reg_sub = ''
						senders[sender] = 0
					
					elif sender in senders and senders[sender] == 2 and resieved_message != 'Нет, ошибка' and resieved_message.lower() != 'нет' and resieved_message != 'Да, всё верно' and resieved_message.lower() != 'да':
						vklog('Пользователь '+str(sender)+', ошибка выбора региона: "'+resieved_message+'"')
						write_message(sender, 'Я вас не понял.&#128530;', 0)
						time.sleep(3)
						senders[sender] = 0
					
					# """ ВВОД РЕГИОНА """
					if sender in senders and senders[sender] == 1 and resieved_message in regions:
						
						vklog('Пользователь '+str(sender)+', предложена подписка, ввод региона: "'+resieved_message+'"')
						
						write_message(sender, 'Вы хотите подписаться на уведомления о прогнозах НМУ по региону\n&#9925;&#9925;&#9925;\n'+regions[resieved_message]+
										'.\n&#9925;&#9925;&#9925;\n Если всё верно, нажмите зеленую кнопку или напишите "да".\n Если допущена ошибка, нажмите красную кнопку или напишите "нет".', 1)
						reg_sub = resieved_message
						senders[sender] = 2
						
					# Если неправильно введен регион
					elif sender in senders and senders[sender] == 1 and resieved_message not in regions:
						
						vklog('Пользователь '+str(sender)+', ошибочный ввод региона: "'+resieved_message+'"')
						write_message(sender, 'Такого региона нет в нашем списке!&#128530;', 0)
						time.sleep(2)
						senders[sender] = 0
			
			
					# """ СПИСОК РЕГИОНОВ, запрос номера региона """
					if sender in senders and senders[sender] == 0:
						# собираем список регионов в строку
						sp_regs = '\n'.join('{}{}'.format(key, val) for key, val in regions.items())
						write_message(sender, 'В настоящий момент доступно информирование о прогнозах НМУ по следующим регионам:'
										+'\n'+'\n'+sp_regs, 0)
						
						write_message(sender,'Если вы хотите подписаться на уведомления о прогнозах НМУ через мессенджер ВКонтакте, отправьте мне номер региона из списка.'+'\n'
										+'Я жду.&#9203;', 0)
						senders[sender] = 1
						
			
	except requests.exceptions.RequestException as e:
		erlog(str(e))
		# постим в тг-канал Error НМУ
		r = requests.post('https://api.telegram.org/bot{0}/sendMessage?chat_id={1}&text={2}{3}{4}'.format(
		t_err, t_err_chatid,
		 'VKBotik: таймаут соединения ВКОНТАКТЕ!',
		 "\n",
		 str(e)))
		continue

	except Exception as e:
		erlog(str(e))
		# постим в тг-канал Error НМУ
		r = requests.post('https://api.telegram.org/bot{0}/sendMessage?chat_id={1}&text={3}{2}{4}'.format(
		t_err, t_err_chatid, "\n",
		 'VKBotik: бот остановлен. Ошибка:',
		str(e)))
		sys.exit()
	
