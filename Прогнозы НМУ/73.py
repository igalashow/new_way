import requests
from datetime import datetime, date, time
import time
import random
import re
import os
from openpyxl import Workbook, load_workbook
from selenium import webdriver
import selenium.common.exceptions
import vk_api
from vk_api.utils import get_random_id
from vk_api.keyboard import VkKeyboard, VkKeyboardColor
from spiski import short_links, ugms, vk_token

reg = '73'				# указать в соответствии с регионом
vecher = 22				# время фиксации прогноза (уже вечер, он не изменится. GMT+3)
url = 'http://pogoda-sv.ru/nmu/index.php'

chatid = ''
api_token = ''

chat_all = ''
api_all = ''

t_err = ""
t_err_chatid = ""

			# ****    ВКОНТАКТЕ   ****
authorize = vk_api.VkApi(token = vk_token)

sub_list = []
if os.path.isfile('/home/cheker/vk/sub'+reg+'.txt'):
	with open('/home/cheker/vk/sub'+reg+'.txt') as f:
		sub_list = f.read().splitlines()

keybo_otpiska = VkKeyboard(one_time=True)
keybo_otpiska.add_button('Отписаться', color=VkKeyboardColor.NEGATIVE)

def write_message(sender, message):	
	authorize.method('messages.send', {'user_id': sender, 'message': message, 'random_id': get_random_id(), 'keyboard': keybo_otpiska.get_keyboard()})
			# ***** конец ВКОНТАКТЕ *****

def erlog(*args):
	l = open("/var/www/html/erlog"+reg+".txt", 'a', encoding='windows-1251')		#в линуксе тут нужен полный путь до файла /var/www/html/NMU_log.txt
	print('Дата проверки '+str(datetime.today())[:-10], file=l)
	print(*args, file=l)
	l.close()

def log(*args):
    with open('/var/www/prognoznmu.ru/nmu'+reg+'_log.txt', encoding='windows-1251') as fr:			# /var/www/prognoznmu.ru/nmu52_log.txt
        tm = fr.read()
    with open('/var/www/prognoznmu.ru/nmu'+reg+'_log.txt', 'w', encoding='windows-1251') as fw:
        print('Дата проверки '+str(datetime.today())[:-10], file=fw)
        print(*args,  tm, sep='\n', file=fw)

def t_error(error):
	r = requests.post('https://api.telegram.org/bot{0}/sendMessage?chat_id={1}&text={3}{2}{4}'.format(
	t_err, t_err_chatid, "\n",
	 'com'+reg+':',
	error,
	'&parse_mode=Markdown&disable_web_page_preview=false'))	# постим в Error НМУ

if not os.path.isfile('/home/cheker/com/re'+reg+'.xlsx'):
	wb = Workbook('/home/cheker/com/re'+reg+'.xlsx')
	ws = wb.create_sheet("Проверка обновления")
	ws.append(["Дата", "Текст"])
	wb.save(os.path.dirname(os.path.abspath(__file__)) + "/re"+reg+".xlsx")
	wb = load_workbook('/home/cheker/com/re'+reg+'.xlsx')
else:
	wb = load_workbook('/home/cheker/com/re'+reg+'.xlsx')
ws = wb.active

while True:

	try:
		#time.sleep(random.randint(0, 300))
		op = webdriver.ChromeOptions()
		op.add_argument('headless')
		op.add_argument('--no-sandbox')
		op.add_argument('--disable-setuid-sandbox')

		driver = webdriver.Chrome(options=op, executable_path=r'/usr/local/bin/chromedriver/chromedriver')

		driver.get(url)
		driver.find_element_by_partial_link_text("Ульяновск").click()		# кликаем на Ульяновск
		
		prognoz = driver.find_element_by_xpath("//div[@id='left_col']")
		st = prognoz.text
		driver.quit()
		
		st = st.replace('Информация о неблагоприятных метеорологических условиях','')
		st = st.replace('Неблагоприятные метеорологические условия','')
		st = st.replace('Рекомендации для населения в периоды НМУ','')
		st_clean = re.sub('[\r\n]', '', st)

		now = datetime.today()
		
		if ws['B2'].value != st_clean:				# проверка обновления прогноза
			#print('Произошло обновление прогноза')
			ws['A2'].value = str(datetime.today())[:-10]
			ws['B2'].value = st_clean
			ws.append([str(datetime.today())[:-10], st_clean])
			wb.save(os.path.dirname(os.path.abspath(__file__)) + "/re"+reg+".xlsx")	#записываем новый прогноз в таблицу

			log('- '+st_clean+"\n")						# логируем извлеченную чистую строку
			
			r = requests.post('https://api.telegram.org/bot{0}/sendMessage?chat_id={1}&text={3}{2}{4}'.format(
				api_token, chatid, "\n", st_clean, 'Источник: ['+ugms[reg]+']('+url+')(кликнуть на Ульяновск)&parse_mode=Markdown&disable_web_page_preview=true'))			# постим в региональный канал

			r = requests.post('https://api.telegram.org/bot{0}/sendMessage?chat_id={1}&text={3}{2}{4}{2}{5}'.format(
				api_all, chat_all, "\n", 'Регион '+reg+' (Ульяновск) LUNAR', st_clean, 'Источник: ['+ugms[reg]+']('+url+')&parse_mode=Markdown&disable_web_page_preview=true'))	# постим в общий техканал

			for sub in sub_list:		# рассылка вконтакте
				time.sleep(0.35)
				write_message(sub, st_clean+'\n\nИсточник: '+ugms[reg]+' '+short_links[reg])


		elif now.hour >= vecher and (now.month != int(ws['A2'].value[5:7]) or now.day != int(ws['A2'].value[8:10])): #не было обновления сегодня и наступил вечер
			ws['A2'].value = str(datetime.today())[:-10]	#обновляем дату
			ws.append([str(datetime.today())[:-10], 'Прогноз НМУ не изменился'])
			wb.save(os.path.dirname(os.path.abspath(__file__)) + "/re"+reg+".xlsx")	#фиксируем отсутствие изменений в таблице

			log('- Прогноз НМУ не изменился'+"\n")						# логируем строку
			
			r = requests.post('https://api.telegram.org/bot{0}/sendMessage?chat_id={1}&text={3}{2}{4}'.format(
				api_token, chatid, "\n", 'Прогноз НМУ не изменился.', 'Источник: ['+ugms[reg]+']('+url+')&parse_mode=Markdown&disable_web_page_preview=true'))			# постим в региональный канал

			r = requests.post('https://api.telegram.org/bot{0}/sendMessage?chat_id={1}&text={3}{2}{4}{2}{5}'.format(
				api_all, chat_all, "\n", 'Регион '+reg+' (Ульяновск)', 'Прогноз НМУ не изменился.', 'Источник: ['+ugms[reg]+']('+url+')&parse_mode=Markdown&disable_web_page_preview=true'))	# постим в общий техканал

			for sub in sub_list:		# рассылка вконтакте
				time.sleep(0.35)
				write_message(sub, 'Прогноз НМУ не изменился.'+'\n\nИсточник: '+ugms[reg]+' '+short_links[reg])

		
	except requests.RequestException as e:
		erlog("OOPS!! General Error")
		erlog(str(e))
		t_error(str(e))		# постинг в ERROR НМУ
	except KeyboardInterrupt:
		erlog(" Кто-то закрыл программу")
		t_error(" Кто-то закрыл программу")		# постинг в ERROR НМУ
	
	except selenium.common.exceptions.NoSuchElementException as e:
		erlog(" Ошибка selenium: элемент не существует")
		erlog(str(e))
		t_error(str(e))		# постинг в ERROR НМУ
	
	except Exception as e:
		erlog(" Ошибка!")
		erlog(str(e))
		t_error(str(e))		# постинг в ERROR НМУ
	
	finally:
		driver.quit()
	break
