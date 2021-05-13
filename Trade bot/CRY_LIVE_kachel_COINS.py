from openpyxl import Workbook, load_workbook
import os
import sys
import urllib, http.client
import time
import requests
import hmac
import hashlib
import datetime
import json
from collections import OrderedDict
import base64
from requests.compat import quote_plus

class CryptopiaApi(object):
    """ Represents a wrapper for cryptopia API """

    def __init__(self, key, secret):
        self.key = key
        self.secret = secret
        self.public = ['GetCurrencies', 'GetTradePairs', 'GetMarkets',
                       'GetMarket', 'GetMarketHistory', 'GetMarketOrders', 'GetMarketOrderGroups']
        self.private = ['GetBalance', 'GetDepositAddress', 'GetOpenOrders',
                        'GetTradeHistory', 'GetTransactions', 'SubmitTrade',
                        'CancelTrade', 'SubmitTip', 'SubmitWithdraw', 'SubmitTransfer']

    def api_query(self, feature_requested, get_parameters=None, post_parameters=None):
        """ Performs a generic api request """
        time.sleep(1)
        if feature_requested in self.private:
            url = "https://www.cryptopia.co.nz/Api/" + feature_requested
            post_data = json.dumps(post_parameters)
            headers = self.secure_headers(url=url, post_data=post_data)
            req = requests.post(url, data=post_data, headers=headers)
            if req.status_code != 200:
                try:
                    req.raise_for_status()
                except requests.exceptions.RequestException as ex:
                    return None, "Status Code : " + str(ex)
            req = req.json()
            if 'Success' in req and req['Success'] is True:
                result = req['Data']
                error = None
            else:
                result = None
                error = req['Error'] if 'Error' in req else 'Unknown Error'
            return (result, error)
        elif feature_requested in self.public:
            url = "https://www.cryptopia.co.nz/Api/" + feature_requested + "/" + \
                  ('/'.join(i for i in get_parameters.values()
                           ) if get_parameters is not None else "")
            req = requests.get(url, params=get_parameters)
            if req.status_code != 200:
                try:
                    req.raise_for_status()
                except requests.exceptions.RequestException as ex:
                    return None, "Status Code : " + str(ex)
            req = req.json()
            if 'Success' in req and req['Success'] is True:
                result = req['Data']
                error = None
            else:
                result = None
                error = req['Error'] if 'Error' in req else 'Unknown Error'
            return (result, error)
        else:
            return None, "Unknown feature"

    def get_currencies(self):
        """ Gets all the currencies """
        return self.api_query(feature_requested='GetCurrencies')

    def get_tradepairs(self):
        """ GEts all the trade pairs """
        return self.api_query(feature_requested='GetTradePairs')

    def get_markets(self):
        """ Gets data for all markets """
        return self.api_query(feature_requested='GetMarkets')

    def get_market(self, market):
        """ Gets market data """
        return self.api_query(feature_requested='GetMarket',
                              get_parameters={'market': market})

    def get_history(self, market):
        """ Gets the full order history for the market (all users) """
        return self.api_query(feature_requested='GetMarketHistory',
                              get_parameters={'market': market})

    def get_orders(self, market, depth):
        """ Gets the user history for the specified market """
        return self.api_query(feature_requested='GetMarketOrders',
                              get_parameters={'market': market, 'depth': str(depth)})

    def get_ordergroups(self, markets):
        """ Gets the order groups for the specified market """
        return self.api_query(feature_requested='GetMarketOrderGroups',
                              get_parameters={'markets': markets})

    def get_balance(self, currency):
        """ Gets the balance of the user in the specified currency """
        result, error = self.api_query(feature_requested='GetBalance',
                                       post_parameters={'Currency': currency})
        if error is None:
            result = result[0]
        return (result, error)

    def get_openorders(self, market):
        """ Gets the open order for the user in the specified market """
        return self.api_query(feature_requested='GetOpenOrders',
                              post_parameters={'Market': market})

    def get_deposit_address(self, currency):
        """ Gets the deposit address for the specified currency """
        return self.api_query(feature_requested='GetDepositAddress',
                              post_parameters={'Currency': currency})

    def get_tradehistory(self, market):
        """ Gets the trade history for a market """
        return self.api_query(feature_requested='GetTradeHistory',
                              post_parameters={'Market': market})

    def get_transactions(self, transaction_type):
        """ Gets all transactions for a user """
        return self.api_query(feature_requested='GetTransactions',
                              post_parameters={'Type': transaction_type})

    def submit_trade(self, market, trade_type, rate, amount):
        """ Submits a trade """
        return self.api_query(feature_requested='SubmitTrade',
                              post_parameters={'Market': market,
                                               'Type': trade_type,
                                               'Rate': rate,
                                               'Amount': amount})

    def cancel_trade(self, trade_type, order_id, tradepair_id):
        """ Cancels an active trade """
        return self.api_query(feature_requested='CancelTrade',
                              post_parameters={'Type': trade_type,
                                               'OrderID': order_id,
                                               'TradePairID': tradepair_id})

    def submit_tip(self, currency, active_users, amount):
        """ Submits a tip """
        return self.api_query(feature_requested='SubmitTip',
                              post_parameters={'Currency': currency,
                                               'ActiveUsers': active_users,
                                               'Amount': amount})

    def submit_withdraw(self, currency, address, amount):
        """ Submits a withdraw request """
        return self.api_query(feature_requested='SubmitWithdraw',
                              post_parameters={'Currency': currency,
                                               'Address': address,
                                               'Amount': amount})

    def submit_transfer(self, currency, username, amount):
        """ Submits a transfer """
        return self.api_query(feature_requested='SubmitTransfer',
                              post_parameters={'Currency': currency,
                                               'Username': username,
                                               'Amount': amount})

    def secure_headers(self, url, post_data):
        """ Creates secure header for cryptopia private api. """
        nonce = str(int(time.time()))
        md5 = hashlib.md5()
        jsonparams = post_data.encode('utf-8')
        md5.update(jsonparams)
        rcb64 = base64.b64encode(md5.digest()).decode('utf-8')
        
        signature = self.key + "POST" + quote_plus(url).lower() + nonce + rcb64
        hmacsignature = base64.b64encode(hmac.new(base64.b64decode(self.secret),
                                                  signature.encode('utf-8'),
                                                  hashlib.sha256).digest())
        header_value = "amx " + self.key + ":" + hmacsignature.decode('utf-8') + ":" + nonce
        return {'Authorization': header_value, 'Content-Type': 'application/json; charset=utf-8'}


class LivecoinAPI(object):

    def get_balance(self, curr):
        request_url = base_url + 'payment/balance'
        data = OrderedDict([('currency', curr)])
        encoded_data = urllib.parse.urlencode(data)
    
        sign = hmac.new(
            secret_key.encode(), msg=encoded_data.encode(),
            digestmod=hashlib.sha256).hexdigest().upper()
    
        headers = {"Api-key": api_key,
                   "Sign": sign}
        ret = requests.get(request_url +'?'+ encoded_data, '', headers=headers)
    
        return ret.json()['value']
    
    def get_all_orderbooks(self, depth):
        request_url = base_url + 'exchange/all/order_book?depth=' + str(depth)
        ret = requests.get(request_url)
        return ret.json()
    
    def get_restrictions(self):
        request_url = base_url + 'exchange/restrictions'
        ret = requests.get(request_url)
        return ret.json()
    
    def get_orderbook(self, currencyPair, depth):
        request_url = base_url + 'exchange/order_book?currencyPair='+currencyPair+'&depth=' + str(depth)
        ret = requests.get(request_url)
        return ret.json()
    
    def sell_market(self, currencyPair, quantity):
        request_url = base_url + 'exchange/sellmarket'
        data = OrderedDict(sorted([('currencyPair', currencyPair),('quantity', quantity)]))
        encoded_data = urllib.parse.urlencode(data)
    
        sign = hmac.new(
            secret_key.encode(), msg=encoded_data.encode(),
            digestmod=hashlib.sha256).hexdigest().upper()
        
        headers = {"Api-key": api_key,
                   "Sign": sign,
                   "Content-type": "application/x-www-form-urlencoded"}
                   
        ret = requests.post(request_url, encoded_data, headers=headers)
        resp = ret.json()
        if resp['success'] == True:
            print('Успешная продажа ', str(quantity), currencyPair[0:-4])
        else:
            print('Продажа {} не удалась!'.format(currencyPair[0:-4]))
        return resp
            
    def buy_market(self, currencyPair, quantity):
        request_url = base_url + 'exchange/buymarket'
        data = OrderedDict(sorted([('currencyPair', currencyPair),('quantity', quantity)]))
        encoded_data = urllib.parse.urlencode(data)
    
        sign = hmac.new(
            secret_key.encode(), msg=encoded_data.encode(),
            digestmod=hashlib.sha256).hexdigest().upper()
        
        headers = {"Api-key": api_key,
                   "Sign": sign,
                   "Content-type": "application/x-www-form-urlencoded"}
        ret = requests.post(request_url, encoded_data, headers=headers)
        resp = ret.json()
        if resp['success'] == True:
            print('Успешная покупка ', str(quantity), currencyPair[0:-4])
        else:
            print('Покупка {} не удалась!'.format(currencyPair[0:-4]))
        return resp
    
    def get_order_info(self, orderId):
        request_url = base_url + 'exchange/order'
        data = OrderedDict([('orderId', orderId)])
        encoded_data = urllib.parse.urlencode(data)
        
        sign = hmac.new(
         secret_key.encode(), msg=encoded_data.encode(),
          digestmod=hashlib.sha256).hexdigest().upper()
        
        headers = {"Api-key": api_key, "Sign": sign}
        
        ret = requests.get(request_url +'?'+ encoded_data, '', headers=headers)
        
        order_i = ret.json()
        #print(order_i)
        return order_i
    
    def sell_limit(self, currencyPair, quantity, price):
        request_url = base_url + 'exchange/selllimit'
        data = OrderedDict(sorted([('currencyPair', currencyPair),('price', price),('quantity', quantity)]))
        encoded_data = urllib.parse.urlencode(data)
        
        sign = hmac.new(
            secret_key.encode(), msg=encoded_data.encode(),
            digestmod=hashlib.sha256).hexdigest().upper()
        
        headers = {"Api-key": api_key,
                   "Sign": sign,
                   "Content-type": "application/x-www-form-urlencoded"}
                   
        ret = requests.post(request_url, encoded_data, headers=headers)
        resp = ret.json()
        if resp['success'] == True:
            print('Успешно создан ордер на продажу', str(quantity), currencyPair[0:-4], 'по курсу', str(price), ' #:', resp['orderId'])
        else:
            print('Ошибка создания ордера на продажу {} !!!'.format(currencyPair[0:-4]))
        return resp
    
    def buy_limit(self, currencyPair, quantity, price):
        request_url = base_url + 'exchange/buylimit'
        data = OrderedDict(sorted([('currencyPair', currencyPair),('price', price),('quantity', quantity)]))
        encoded_data = urllib.parse.urlencode(data)
        
        sign = hmac.new(
            secret_key.encode(), msg=encoded_data.encode(),
            digestmod=hashlib.sha256).hexdigest().upper()
        
        headers = {"Api-key": api_key,
                   "Sign": sign,
                   "Content-type": "application/x-www-form-urlencoded"}
                   
        ret = requests.post(request_url, encoded_data, headers=headers)
        resp = ret.json()
        if resp['success'] == True:
            print('Успешно создан ордер на покупку', str(quantity), currencyPair[0:-4], 'по курсу', str(price), ' #:', resp['orderId'])
        else:
            print('Ошибка создания ордера на покупку {} !!!'.format(currencyPair[0:-4]))
        return resp
    
    def cancel_order(self, currencyPair, orderId):
        request_url = base_url + 'exchange/cancellimit'
        data = OrderedDict(sorted([('currencyPair', currencyPair), ('orderId', orderId)]))
        encoded_data = urllib.parse.urlencode(data)
        
        sign = hmac.new(
            secret_key.encode(), msg=encoded_data.encode(),
            digestmod=hashlib.sha256).hexdigest().upper()
        
        headers = {"Api-key": api_key,
                   "Sign": sign,
                   "Content-type": "application/x-www-form-urlencoded"}
                   
        ret = requests.post(request_url, encoded_data, headers=headers)
        resp = ret.json()
        if resp['cancelled'] == True:
            print('Успешно отменен ордер {} {}'.format(orderId, currencyPair))
        else:
            print('Ошибка отмены ордера {} !!!'.format(orderId))
        return resp
    
    def get_orders(self, openClosed='OPEN'):
        request_url = base_url + 'exchange/client_orders'
        data = OrderedDict([('openClosed', openClosed)])
        encoded_data = urllib.parse.urlencode(data)
        
        sign = hmac.new(
         secret_key.encode(), msg=encoded_data.encode(),
          digestmod=hashlib.sha256).hexdigest().upper()
        
        headers = {"Api-key": api_key, "Sign": sign}
        
        ret = requests.get(request_url +'?'+ encoded_data, '', headers=headers)
        
        return ret.json()

key = '6be7fd6bab224f308af72e22c051f1da'
secret = 'AWu14f7RxMOXO29dUgKG7Oi3kXWSCgDC0XOoY3wRms4='     # CRYPTOPIA

base_url = 'https://api.livecoin.net/'                      # LIVECOIN
api_key = '8SBx7hB3j31XBZQPgXD94dbZmbaXkEgd'
secret_key = 'VyXhBZ2jj3wHY8RhVaKGMqs8WYtPsukQ'

birga1 = 'CRYP'
birga2 = 'LIVE'
fee1 = 0.002
fee2 = 0.0018
lot = 0.00051 # Размер лота
perc = 0.01     # Сколько процентов хотим заработать минимум
tradecoin = 'BTC'       # Торговая монета
COINS = ['DBIX']            # Оборотные монеты
USE_LOG = True

def log(*args):
	""" Логирование работы бота """
    if USE_LOG:
       l = open("./crypto_live_kachelCOINS.txt", 'a', encoding='utf-8')
       print(datetime.datetime.now(), *args, file=l)
       l.close()
    print(datetime.datetime.now(),' ', *args)


if not os.path.isfile('./crypto_live_kachelCOINS.xlsx'):
    wb = Workbook('crypto_live_kachelCOINS.xlsx')
    ws = wb.create_sheet("checker")
    ws.append(["Дата сделки", "Пара сделки", "Покупка", "Продажа", "Профит", "% профита"])
    wb.save(os.path.dirname(os.path.abspath(__file__)) + "/crypto_live_kachelCOINS.xlsx")
    wb = load_workbook('crypto_live_kachelCOINS.xlsx')
else:
    wb = load_workbook('crypto_live_kachelCOINS.xlsx')
ws = wb.active

l = LivecoinAPI()
c = CryptopiaApi(key, secret)

while True:
    
    try: 
        for coin in COINS:
            print(datetime.datetime.now(), 'Проверка ',coin)
            # while l.get_balance(tradecoin) < lot and c.get_balance(tradecoin)[0]['Available'] < lot:
                # log('Нет {} ни на одной бирже, ждем поступления...'.format(tradecoin))
                # time.sleep(30)
    
            timezap = time.time()
            print(datetime.datetime.now(),' Запрашиваем данные с бирж...')
            
            try:
                crypto_resp = c.get_orders(market=coin+'_'+tradecoin, depth=1)
            except Exception:
                log('Что-то не так с соединением CRYPTOPIA. Повторное подключение...')
                time.sleep(10)
                continue
            
            try:
                live_resp = l.get_orderbook(currencyPair=coin+'/'+tradecoin, depth=1)
            except Exception:
                log('Что-то не так с соединением LIVECOIN. Повторное подключение...')
                time.sleep(10)
                continue
            
            lag = round(time.time()-timezap)
            print(datetime.datetime.now(),' Данные получены, затрачено ', lag, 'сек.')
            if lag <= 4:
                
                if not crypto_resp[0]['Sell']:
                    # Нет ордеров на продажу CRYPTOPIA
                    ask1 = -1
                    vol_ask1 = -1
                else:
                    ask1 = float(crypto_resp[0]['Sell'][0]['Price'])
                    vol_ask1 = float(crypto_resp[0]['Sell'][0]['Volume'])
                if not crypto_resp[0]['Buy']:
                    # Нет ордеров на покупку CRYPTOPIA
                    bid1 = -1
                    vol_bid1 = -1
                else:
                    bid1 = float(crypto_resp[0]['Buy'][0]['Price'])
                    vol_bid1 = float(crypto_resp[0]['Buy'][0]['Volume'])
                
                
                if not live_resp['asks']:
                    # Нет ордеров на продажу LIVECOIN
                    ask2 = -1
                    vol_ask2 = -1
                else:
                    ask2 = float(live_resp['asks'][0][0])
                    vol_ask2 = float(live_resp['asks'][0][1])
                if not live_resp['bids']:
                    # Нет ордеров на покупку LIVECOIN
                    bid2 = -1
                    vol_bid2 = -1
                else:
                    bid2 = float(live_resp['bids'][0][0])
                    vol_bid2 = float(live_resp['bids'][0][1])
                
                # Результат покупки на бирже 1
                prom_birga1 = round(lot/ask1*(1-fee1), 8)
                # Результат покупки на бирже 2
                prom_birga2 = round(lot/ask2*(1-fee2), 8)
                # Итоговая сумма на бирже 1 (покупали на бирже 2)
                itog_birga1 = round(prom_birga2*bid1*(1-fee1), 8)
                # Итоговая сумма на бирже 2 (покупали на бирже 1)
                itog_birga2 = round(prom_birga1*bid2*(1-fee2), 8)
                
                # Вариант 2 CRY-LIVE
                if itog_birga2 > lot*(1+perc) and vol_ask1 >= lot/ask1 and vol_bid2 >= lot/ask1:
                    b1_t = c.get_balance(tradecoin)[0]['Available']
                    b2_c = l.get_balance(coin)
                    
                    # если балансы в порядке
                    if b1_t >= lot and b2_c >= prom_birga1:     
                        part = False
                        cancel1 = {'exception': ''}
                        log('{} {}: продаем {} {}, цена {}, получим {} {}'.format(birga2, coin+'/'+tradecoin,
                                prom_birga1, coin, bid2, itog_birga2, tradecoin))
                        
                        # Продаем на LIVE
                        sdelka1 = l.sell_limit(currencyPair=coin+'/'+tradecoin, price=bid2, quantity=prom_birga1)       
                        log(sdelka1)
                        
                        # выставлен лимитный ордер
                        if sdelka1['success'] == True and sdelka1['added'] == True:
                            i = l.get_order_info(orderId=sdelka1['orderId'])
                            
                            # если нет частичного исполнения
                            if i['quantity'] == i['remaining_quantity']:  
                                log('Не успели! Отменяем выставленный ордер на продажу {} {}'.format(prom_birga1, coin))
                                cancel1 = l.cancel_order(currencyPair=coin+'/'+tradecoin, orderId=sdelka1['orderId'])
                                log(cancel1)
                                if cancel1['exception'] == 'Cannot find order':
                                    log('Ордер все же сработал, но это даже к лучшему: продолжаем...')
                            
                            # если частичное исполнение
                            else:                                           
                                part = True
                                log('Частичное исполнение ордера. Переходим ко второй сделке, надеемся на полное исполнение.')
                        
                        # Если успешная сделка на LIVE
                        if (sdelka1['success'] == True and sdelka1['added'] == False) or cancel1['exception'] == 'Cannot find order' or part == True:   
                            log('{} {}: покупаем {}, цена {}, получим {} {}'.format(birga1, coin+'/'+tradecoin, 
                            coin, ask1, prom_birga1, coin))
                            
                            # Покупаем на CRY
                            sdelka2 = c.submit_trade(market=coin+'_'+tradecoin, trade_type='Buy', rate=ask1, amount=prom_birga1)    
                            log(sdelka2)
                            
                            # Если выставился лимитный ордер
                            if sdelka2[0]['OrderId'] != None:               
                                log('Не успели. {} {} зависли на лимитном ордере ({}). Предполагаемая прибыль помечена *'.format(lot, tradecoin, birga1))
                                # log('Ждем срабатывания ордера...')
                                # while c.get_balance(currency=coin)[0]['Available'] < prom_birga1:
                                    # time.sleep(30)
                                # log('Баланс '+coin+' больше лота. Продолжаем.')
                                ws.append([datetime.datetime.now(),
                                            coin+'/'+tradecoin, birga1, birga2, itog_birga2-lot,
                                            round(itog_birga2*100/lot-100, 2), '*'])
                                wb.save(os.path.dirname(os.path.abspath(__file__)) + "/crypto_live_kachelCOINS.xlsx")
                            
                            # Успешный арбитраж по варианту 2
                            if sdelka2[0]['FilledOrders'] and sdelka2[0]['OrderId'] == None:          
                                log('Успешный арбитраж! Прибыль {} {}, {}% (зафиксирована на {})'.format(
                                    itog_birga2-lot, tradecoin, round(itog_birga2*100/lot-100, 2), birga2))
                                ws.append([datetime.datetime.now(),
                                            coin+'/'+tradecoin, birga1, birga2, itog_birga2-lot,
                                            round(itog_birga2*100/lot-100, 2)])
                                wb.save(os.path.dirname(os.path.abspath(__file__)) + "/crypto_live_kachelCOINS.xlsx")

                            # else:
                                # log('Что-то пошло не так, смотри логи. Выход')
                                # sys.exit()       
                        # else:
                            # log('Что-то пошло не так, смотри логи. Выход')
                            # sys.exit()
                    else:
                        log('Нет достаточных средств на биржах:{} {} {}, {} {} {}'.format(
                            birga1, b1_t, tradecoin, birga2, b2_c, coin))
                                
                        
               
               
                # Вариант 1 LIVE-CRY
                elif itog_birga1 > lot*(1+perc) and vol_ask2 >= lot/ask2 and vol_bid1 >= lot/ask2:
                    b2_t = l.get_balance(tradecoin)
                    b1_c = c.get_balance(coin)[0]['Available']
                    
                    # если балансы в порядке
                    if b2_t >= lot and b1_c >= prom_birga2:
                        cancel2 = [1]
                        log('{} {}: продаем {} {}, цена {}, получим {} {}'.format(birga1, coin+'/'+tradecoin,       
                                prom_birga2, coin, bid1, itog_birga1, tradecoin))
                        # продаем на CRYPTOPIA
                        sdelka1 = c.submit_trade(market=coin+'_'+tradecoin, trade_type='Sell', rate=bid1, amount=prom_birga2)
                        log(sdelka1)
                        
                        # Если выставлен лимитный ордер в полном объеме
                        if not sdelka1[0]['FilledOrders'] and sdelka1[0]['OrderId'] != None:            
                            log('Не успели! Отменяем выставленный ордер на продажу {} {}'.format(prom_birga2, coin))
                            cancel2 = c.cancel_trade(trade_type='Trade', order_id=sdelka1[0]['OrderId'], tradepair_id='')
                            log(cancel2)
                            if cancel2[0] == None:
                                log('Ордер все же сработал, но это даже к лучшему: продолжаем...')
                        
                        # Если успешная сделка на CRY
                        if sdelka1[0]['FilledOrders'] or cancel1[0] == None:
                            # подвариант: частично исполненный ордер
                            if sdelka1[0]['OrderId'] != None:
                                log('Частичное исполнение ордера. Переходим ко второй сделке, надеемся на полное исполнение.')
                                
                            log('{} {}: покупаем {}, цена {}, получим {} {}'.format(birga2, coin+'/'+tradecoin, 
                                    coin, ask2, prom_birga2, coin))
                            
                            # Покупаем на LIVE
                            sdelka2 = l.buy_limit(currencyPair=coin+'/'+tradecoin, price=ask2, quantity=prom_birga2)          
                            log(sdelka2)
                            
                            # Если выставлен лимитный ордер
                            if sdelka2['success'] == True and sdelka2['added'] == True:     
                                log('Не успели. {} {} зависли на лимитном ордере ({}). Предполагаемая прибыль помечена *'.format(lot, tradecoin, birga2))  
                                # while l.get_balance(curr=coin) < prom_birga2:
                                    # time.sleep(30)
                                # log('Баланс '+coin+' больше лота. Продолжаем.')
                                ws.append([datetime.datetime.now(), coin+'/'+tradecoin, birga2, birga1, itog_birga1-lot,
                                            round(itog_birga1*100/lot-100, 2), '*'])  
                                wb.save(os.path.dirname(os.path.abspath(__file__)) + "/crypto_live_kachelCOINS.xlsx")
                            
                            # Успешный арбитраж по варианту 1
                            elif sdelka2['success'] == True and sdelka2['added'] == False:                      
                                log('Успешный арбитраж! Прибыль {} {}, {}% (зафиксирована на {})'.format(
                                    itog_birga1-lot, tradecoin, round(itog_birga1*100/lot-100, 2), birga1))
                                ws.append([datetime.datetime.now(), coin+'/'+tradecoin, birga2, birga1, itog_birga1-lot,
                                            round(itog_birga1*100/lot-100, 2)])  
                                wb.save(os.path.dirname(os.path.abspath(__file__)) + "/crypto_live_kachelCOINS.xlsx")
                            # else:
                                # log('Что-то пошло не так, смотри логи. Выход')
                                # sys.exit()
                        # else:
                            # log('Что-то пошло не так, смотри логи. Выход')
                            # sys.exit()
                    else:
                        log('Нет достаточных средств на биржах:{} {} {}, {} {} {}'.format(
                            birga1, b1_c, coin, birga2, b2_t, tradecoin))
                        
            
            
            else:
                print('Слишком большой лаг - {} сек, нужны более актуальные данные'.format(lag))
            time.sleep(2)

    except Exception as e:
        log(e)
