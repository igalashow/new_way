import os
from datetime import datetime
import telebot
import yadisk
from openpyxl import Workbook, load_workbook
from journalist_conf import TOKEN, CHATID, YA_TOKEN

bot = telebot.TeleBot(TOKEN)
y = yadisk.YaDisk(token=YA_TOKEN)
patch = ''  # путь к файлу

# пишем старый файл в список
with open(patch + 'journalsGNSS.txt') as f:
    open_journal: str = f.read().splitlines()[0]


@bot.message_handler(commands=['хелп'])
def help_command(message):
    """ Список команд бота """
    bot.send_message(
        CHATID,
        'Список команд: \n' +
        '🌐*/журнал* <объект> - завести новый журнал по объекту\n' +
        '🌐*/запись* <пункт> <приёмник> <высота> <тип высоты> - запись о сеансе\n' +
        '🌐*/скачать* - скачать журнал в формате XLSX из списка',
        parse_mode="Markdown"
    )


@bot.message_handler(content_types=['text'])
def work_command(message):
    """ Отработка рабочих команд """
    timemark = datetime.now().strftime('%d.%m.%y %H-%M-%S')

    if '/журнал' in message.text:
        obj = message.text[8:]
        j_name = obj + ' ' + timemark
        bot.send_message(CHATID, "📗Завожу новый журнал: " + j_name)

        global open_journal
        old_journal = open_journal
        open_journal = j_name + '.xlsx'

        # Cоздаем новый журнал
        wb = Workbook(patch + open_journal)
        ws = wb.create_sheet(timemark)
        ws.append(["Пункт", "Приёмник", "Высота", "Тип высоты"])
        wb.save(os.path.dirname(os.path.abspath(__file__)) + '/' + open_journal)

        # пишем новый журнал в список
        f = open(patch + 'journalsGNSS.txt', 'w')
        f.write(open_journal)
        f.close()

        bot.send_message(CHATID, "Пиши: /запись <пункт> <приёмник> <высота> <тип высоты>")

        # удаляем старый журнал
        os.remove(patch + old_journal)

    if '/запись' in message.text:
        record = message.text[8:].split(sep=' ')  # записи в список

        wb = load_workbook(patch + open_journal)  # записываем в эксель журнал данные
        ws = wb.active
        ws.append(rec for rec in record)
        wb.save(os.path.dirname(os.path.abspath(__file__)) + '/' + open_journal)

        try:
            # если есть файл и есть олд
            if y.exists('/GNSS/' + open_journal) and y.exists('/GNSS/' + open_journal + '_old'):
                # удаляем олд
                y.remove('/GNSS/' + open_journal + '_old')  
                # переименовываем в олд
                y.move('/GNSS/' + open_journal, '/GNSS/' + open_journal + '_old')
                # Загружаем файл на ядиск
                y.upload(patch + open_journal, '/GNSS/' + open_journal)

            # если есть файл на ядиске, но нет олда
            elif y.exists('/GNSS/' + open_journal) and not y.exists('/GNSS/' + open_journal + '_old'):
                # переименовываем в олд
                y.move('/GNSS/' + open_journal, '/GNSS/' + open_journal + '_old')
                # Загружаем файл на ядиск
                y.upload(patch + open_journal, '/GNSS/' + open_journal)
            
            # если нет файла на диске
            else:
                # Загружаем файл на ядиск
                y.upload(patch + open_journal, '/GNSS/' + open_journal)  

        except Exception as e:
            bot.send_message(CHATID, "Ошибка:\n " + str(e))

        bot.send_message(CHATID, "Записано в журнал 📗 " + open_journal)

    if '/скачать' in message.text:
        bot.send_message(CHATID, "Функция в разработке")


try:
    bot.polling(none_stop=True, interval=1)
except:
    pass
